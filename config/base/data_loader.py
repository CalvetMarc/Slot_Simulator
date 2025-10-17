import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

def load_game_tables(table_names, base_path, file_name="slot_config.xlsx"):
    """
    Loads specific tables (Grid, Strips, Paylines, Paytable...) from the Excel file.
    Supports both true Excel Tables and plain sheet names (Google Sheets exports).
    Returns a dict {table_name_lower: DataFrame}.
    """
    excel_file = Path(base_path) / file_name
    wb = load_workbook(excel_file, data_only=True)
    result = {}    

    for table_name in table_names:
        df = None
        found_table = False

        # 1️⃣ Try to load as real Excel Table (with .ref)
        for ws in wb.worksheets:
            if table_name in ws.tables:
                table_obj = ws.tables[table_name]

                # Comprovem si és un objecte amb .ref
                if hasattr(table_obj, "ref"):
                    ref = table_obj.ref
                    start_cell, end_cell = ref.split(":")
                    start_col = ''.join([c for c in start_cell if c.isalpha()])
                    start_row = int(''.join([c for c in start_cell if c.isdigit()]))
                    end_col = ''.join([c for c in end_cell if c.isalpha()])
                    end_row = int(''.join([c for c in end_cell if c.isdigit()]))

                    df = pd.read_excel(
                        excel_file,
                        sheet_name=ws.title,
                        header=None,
                        usecols=f"{start_col}:{end_col}",
                        skiprows=start_row - 1,
                        nrows=end_row - start_row + 1
                    )

                    df.columns = df.iloc[0]
                    df = df[1:].reset_index(drop=True)
                    result[table_name.lower()] = df
                    found_table = True
                    break

        # 2️⃣ Fallback: try to load by sheet name
        if not found_table:
            if table_name in wb.sheetnames:
                print(f"📄 Loading sheet '{table_name}' directly (no Excel Table found).")
                df = pd.read_excel(excel_file, sheet_name=table_name)
                result[table_name.lower()] = df
            else:
                print(f"⚠️ Table or sheet '{table_name}' not found in workbook.")

    print(f"Loaded tables: {list(result.keys())} ✅")
    return result
